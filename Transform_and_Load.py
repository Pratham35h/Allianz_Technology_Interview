import os
import shutil
import zipfile
from bs4 import BeautifulSoup
from openpyxl import Workbook

global max_year
global min_year

def rename_and_zip_files(directory, zip_filename):
    if not os.path.exists(directory):
        print(f"Directory '{directory}' does not exist.")
        return

    renamed_dir = os.path.join("transformed", "renamed_files")
    if not os.path.exists(renamed_dir):
        os.makedirs(renamed_dir)

    file_list = [f for f in os.listdir(
        directory) if os.path.isfile(os.path.join(directory, f))]
    for filename in file_list:
        original_path = os.path.join(directory, filename)
        temp_copy_path = os.path.join(renamed_dir, filename)
        shutil.copy2(original_path, temp_copy_path)

    temp_file_list = os.listdir(renamed_dir)
    for index, filename in enumerate(temp_file_list, start=1):
        file_extension = os.path.splitext(filename)[1]
        new_name = f"{index}{file_extension}"
        original_path = os.path.join(renamed_dir, filename)
        renamed_path = os.path.join(renamed_dir, new_name)
        os.rename(original_path, renamed_path)

    zip_path = os.path.join("transformed", zip_filename)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in os.listdir(renamed_dir):
            zipf.write(os.path.join(renamed_dir, file), arcname=file)

    shutil.rmtree(renamed_dir)

    # print(f"Files renamed and zipped into '{zip_path}' successfully.")

def natural_sort_key(filename):
    base_name = os.path.splitext(filename)[0]  # Remove the file extension
    try:
        return int(base_name.split("_")[1])
    except (IndexError, ValueError):
        return float('inf')

def create_nhl_stats_sheet(source_directory, output_file):
    all_rows = []
    years = []
    header_added = False

    for filename in sorted(os.listdir(source_directory), key = natural_sort_key):
        if filename.endswith(".html"):
            file_path = os.path.join(source_directory, filename)

            with open(file_path, "r", encoding="utf-8") as file:
                soup = BeautifulSoup(file, "html.parser")

            table = soup.find("table", {"class": "table"})
            if table:
                rows = table.find_all("tr")
                for index, row in enumerate(rows):
                    columns = [col.text.strip() for col in row.find_all(["th", "td"])]
                    
                    if index == 0 and not header_added:
                        all_rows.append(columns)
                        header_added = True  # Ensure the header is added only once
                    elif index > 0:
                        all_rows.append(columns)
                
                year_col_index = all_rows[0].index("Year")
                years.extend(int(row[year_col_index]) for row in all_rows[1:] if row[year_col_index].isdigit())

    if years:
        global min_year
        global max_year
        min_year = min(years)
        max_year = max(years)
        sheet_title = f"NHL Stats {min_year}-{max_year}"
        output_file = os.path.join(output_file,f"NHL_Stats_{min_year}-{max_year}.xlsx")
    else:
        sheet_title = "NHL Stats"
        output_file = os.path.join(output_file,"NHL_Stats.xlsx")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title

    for row in all_rows:
        sheet.append(row)
    
    create_winner_loser_summary_from_array(workbook)

    workbook.save(output_file)
    print(f"Excel file created successfully at '{output_file}'.")

def create_team_year_wins_array(workbook):
    global min_year
    global max_year
    sheet = workbook[f"NHL Stats {min_year}-{max_year}"]
    headers = [cell.value for cell in sheet[1]]
    # print(headers)
    year_col = headers.index("Year")
    team_col = headers.index("Team Name")
    wins_col = headers.index("Wins")

    team_year_wins = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        year = row[year_col]
        team = row[team_col]
        wins = row[wins_col]

        if year and team and wins:
            team_year_wins.append([team, int(year), int(wins)])

    return team_year_wins

def create_winner_loser_summary_from_array(workbook):
    team_year_wins = create_team_year_wins_array(workbook)
    summary_data = {}

    for team, year, wins in team_year_wins:
        if year not in summary_data:
            summary_data[year] = {
                "Winner": (team, wins),
                "Loser": (team, wins)
            }
        else:
            if wins > summary_data[year]["Winner"][1]:
                summary_data[year]["Winner"] = (team, wins)
            if wins < summary_data[year]["Loser"][1]:
                summary_data[year]["Loser"] = (team, wins)

    sheet = workbook.create_sheet("Winner and Loser per Year")
    sheet.append(["Year", "Winner", "Winner Num. of Wins", "Loser", "Loser Num. of Wins"])  # Header

    for year, data in sorted(summary_data.items()):
        sheet.append([
            year,
            data["Winner"][0], data["Winner"][1], 
            data["Loser"][0], data["Loser"][1]
        ])

    # print("Sheet 'Winner and Loser per Year' created successfully.")
